from TDhelper.generic.transformationType import transformation
from TDhelper.document.excel.FieldType import *
from types import FunctionType, MethodType, ModuleType
from openpyxl import load_workbook
import copy

class _AttributeOverride:
    def __init__(self, name, m_type):
        self._name = name
        self._type = m_type

    def __get__(self, instance, owen):
        return instance.__dict__[self._name]

    def __set__(self, instance, value):
        instance.__dict__[self._name] = transformation(value, self._type)

    def __delete__(self, instance):
        instance.__dict__.pop(self._name)


class modelMeta(type):
    def __new__(cls, name, bases, dct):
        attrs = {'mapping': {}, 'Meta': Meta,
                 '__exit__': __exit__, 'readLine': readLine, 'close': close}
        for name, value in dct.items():
            if (not isinstance(dct[name], type)) and (not isinstance(dct[name], FunctionType)):
                if not name.startswith('__'):
                    if isinstance(dct[name], FieldType):
                        attrs['mapping'][name] = value.bindCol
                        attrs[name] = _AttributeOverride(name, value.fieldType)
                    else:
                        raise Exception('field type must is FieldType.')
            else:
                if isinstance(dct[name], type):
                    if name == 'Meta':
                        for attr_name in dct[name].__dict__:
                            if not attr_name.startswith('__'):
                                setattr(attrs['Meta'], attr_name,
                                        dct[name].__dict__[attr_name])
                else:
                    attrs[name] = value
        try:
            if 'Meta' in attrs:
                if attrs['Meta'].file:
                    attrs['excelHandle'] = load_workbook(attrs['Meta'].file)
                    if attrs['excelHandle']:
                        attrs['sheetHandle'] = attrs['excelHandle'][attrs['Meta'].sheet]
                    else:
                        raise Exception(
                            'open ''{0}'' excel file error.' % attrs['Meta'].file)
        except Exception as e:
            raise e
        return super(modelMeta, cls).__new__(cls, name, bases, attrs)


class Meta:
    file = None
    sheet = 'sheet1'


def __exit__(self, exc_type, exc_value, exc_t):
    self.close()

def close(self):
    self.excelHandle = None
    self.sheetHandle = None

def readLine(self, lineOffset=1):
    if lineOffset < 1:
        lineOffset = 1
    if self.sheetHandle:
        if lineOffset <= self.sheetHandle.max_row:
            rowdata = []
            column = self.sheetHandle.max_column+1
            for i in range(1, column):
                cellvalue = self.sheetHandle.cell(
                    row=lineOffset, column=i).value
                rowdata.append(cellvalue)
            for (name, value) in self.mapping.items():
                if value <= len(rowdata):
                    setattr(self, name, rowdata[value-1])
                else:
                    raise Exception('mapping error:(%s,%s)' % (name, value))
            return self
        else:
            return None
    else:
        raise Exception('Sheet Handle is None.')
