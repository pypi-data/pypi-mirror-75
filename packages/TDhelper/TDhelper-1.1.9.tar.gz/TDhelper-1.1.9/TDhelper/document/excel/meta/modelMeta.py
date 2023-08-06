from TDhelper.generic.transformationType import transformation
from FieldType import *
from types import FunctionType, MethodType, ModuleType
from openpyxl import load_workbook
import copy


class _AttributeOverride:
    def __init__(self, name, m_type):
        self._name = name
        self._type = m_type

    def __get__(self, instance, owen):
        return instance.__dict__[self._name]

    def __set__(self, instance, value):
        instance.__dict__[self._name] = transformation(value, self._type)

    def __delete__(self, instance):
        instance.__dict__.pop(self._name)

class Meta:
    file = None
    sheet = 'sheet1'


class modelMeta(type):
    def __new__(cls, name, bases, dct):
        attrs = {'mapping': {}, 'Meta':Meta, '__exit__': __exit__, 'readLine': readLine,
                 'close': close, '__initExcelHandle__': __initExcelHandle__}
        for name, value in dct.items():
            if (not isinstance(dct[name], type)) and (not isinstance(dct[name], FunctionType)):
                if not name.startswith('__'):
                    if isinstance(dct[name], FieldType):
                        attrs['mapping'][name] = value.bindCol
                        attrs[name] = _AttributeOverride(name, value.fieldType)
                    else:
                        raise Exception('field type must is FieldType.')
            else:
                if isinstance(dct[name], type):
                    if name == 'Meta':
                        for attr_name in dct[name].__dict__:
                            if not attr_name.startswith('__'):
                                setattr(attrs['Meta'], attr_name, dct[name].__dict__[attr_name])
                else:
                    attrs[name] = value
        return super(modelMeta, cls).__new__(cls, name, bases, attrs)




def __initExcelHandle__(self):
    try:
        if self.Meta.file:
            self.__excelHandle__ = load_workbook(self.Meta.file)
            self.__sheetHandle__ = self.__excelHandle__[self.Meta.sheet]
        else:
            raise Exception('meta file is None.')
    except Exception as e:
        raise e


def __exit__(self, exc_type, exc_value, exc_t):
    self.close()


def close(self):
    self.__excelHandle__ = None
    self.__sheetHandle__ = None


def readLine(self, lineOffset=1):
    if lineOffset < 1:
        lineOffset = 1
    if self.__sheetHandle__:
        if lineOffset <= self.__sheetHandle__.max_row:
            rowdata = []
            column = self.__sheetHandle__.max_column+1
            for i in range(1, column):
                cellvalue = self.__sheetHandle__.cell(
                    row=lineOffset, column=i).value
                rowdata.append(cellvalue)
            for (name, value) in self.mapping.items():
                if value <= len(rowdata):
                    setattr(self, name, rowdata[value-1])
                else:
                    raise Exception('mapping error:(%s,%s)' % (name, value))
            return self
        else:
            return None
    else:
        raise Exception('Sheet Handle is None.')
