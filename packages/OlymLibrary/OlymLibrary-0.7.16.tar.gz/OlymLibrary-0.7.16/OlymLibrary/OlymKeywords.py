#-*- coding:utf-8 -*-
__version__='0.1'
from robot.api import logger
from selenium import webdriver
from time import sleep
import re
import datetime,time
import random
import json
import types
import os
import webbrowser
from robot.api import logger
# import xlwt  # 俞森 2018-12-14   写入excel时用
# from xlutils.copy import copy
# from xlrd import *  # 俞森 2018-12-14   写入excel时用
# import xlrd
import sys
import hashlib  # 俞森 2019-04-03  进行md5加密时用
import urllib.request, urllib.parse, urllib.error
import collections
#import imp
import openpyxl  # 俞森 2020-07-27  python3写入excel相关
from openpyxl.styles import Font  # 俞森 2020-07-27  python3写入excel相关

#imp.reload(sys)
#sys.setdefaultencoding( "utf-8" )

class OlymKeywords(object):
    def XlsToJson(self, filePath, sheetname='Sheet1', x1=0, x2=0):
        reg = re.compile('^(\w*)$')
        regArr1 = re.compile(r'''
    		^(\w*)
    		\[(\d*)\]
    		_(\w*)$
    		''', re.VERBOSE)
        regArr2 = re.compile(r'''
    		^(\w*)
    		\[(\d*)\]
    		<(\w*)>$''', re.VERBOSE)
        # regArr3 = re.compile(r'''
        # 	^(\w*)
        # 	\[(\d*)\]
        # 	_(\w*)$
        # ''',re.VERBOSE)
        nType = 0
        # print( 'filePath: ' + filePath + ' savePath: ' + savePath )
        data = xlrd.open_workbook(filePath)
        table = data.sheet_by_name(sheetname)
        commentInfo = []
        headerInfo = []
        xlsInfo = {}
        ids = []
        if x1 < 3 or x1 > table.nrows:
            x1 = 3
        if x2 < 2 or x2 > table.nrows:
            x2 = table.nrows

        for i in range(table.ncols):
            l = 2
            commentInfo.append(table.cell(0, i).value)
            header = table.cell(1, i).value
            headerInfo.append(header)
            match = regArr2.match(header)
            if match:
                nType = 3
            else:
                match = regArr1.match(header)
                if match:
                    nType = 2
                else:
                    match = reg.match(str(header))
                    # print(match.group(0))
                    nType = 1
            nId = 0
            for j in range(x1 - 1, x2):
                cellInfo = table.cell(j, i).value
                # 判断excel读取的元素的类型
                type = table.cell(j, i).ctype
                if str(cellInfo) == 'true':
                    cellInfo = True
                if str(cellInfo) == 'false':
                    cellInfo = False
                if type == 2 and cellInfo % 1 == 0:
                    cellInfo = int(cellInfo)
                elif type == 4:
                    cell = True if cell == 1 else False
                # end
                if i == 0:
                    # cellInfo = int(cellInfo)
                    ids.append(l - 1)
                    xlsInfo[l - 1] = {}
                    nType = 0
                nId = int(ids[l - 2])
                l += 1
                if nType == 0:
                    # if match.group(1) not in xlsInfo[nId] and reg.match(str(header))==False:
                    # 	xlsInfo[nId][match.group(1)] = {}
                    # 	xlsInfo[nId][match.group(1)][match.group(3)] = cellInfo
                    # if match.group(1) not in xlsInfo[nId] and reg.match(str(header))==False and regArr1.match(header):
                    if match.group(1) not in xlsInfo[nId] and regArr1.match(header):
                        xlsInfo[nId][match.group(1)] = {}
                        xlsInfo[nId][match.group(1)][match.group(3)] = cellInfo
                    else:
                        xlsInfo[nId][match.group(1)] = cellInfo
                if nType == 1:
                    xlsInfo[nId][match.group(1)] = cellInfo
                elif nType == 2:
                    if match.group(1) not in xlsInfo[nId]:
                        xlsInfo[nId][match.group(1)] = {}
                    xlsInfo[nId][match.group(1)][match.group(3)] = cellInfo
                elif nType == 3:
                    if match.group(1) not in xlsInfo[nId]:
                        xlsInfo[nId][match.group(1)] = []
                    if len(xlsInfo[nId][match.group(1)]) < int(match.group(2)):
                        xlsInfo[nId][match.group(1)].append({})
                    xlsInfo[nId][match.group(1)][int(match.group(2)) - 1][match.group(1)] = cellInfo

        # outFile = codecs.open( savePath, 'w', 'utf-8' )
        listinfo = []

        # for k in range(len(listinfo)):
        #	del listinfo[k]['ID']

        for key, value in list(xlsInfo.items()):
            listinfo.append(value)
        # print('序号:'+str(key)+'  json数据:'+str(value))

        for k in range(len(listinfo)):
            #del listinfo[k]['ID']
            listinfo[k] = json.dumps(listinfo[k])
            print((listinfo[k]))

        return listinfo

    def XlsToDic(self, filePath, sheetname='Sheet1', x1=0, x2=0):
        reg = re.compile('^(\w*)$')
        regArr1 = re.compile(r'''
    		^(\w*)
    		\[(\d*)\]
    		_(\w*)$
    		''', re.VERBOSE)
        regArr2 = re.compile(r'''
    		^(\w*)
    		\[(\d*)\]
    		<(\w*)>$''', re.VERBOSE)
        # regArr3 = re.compile(r'''
        # 	^(\w*)
        # 	\[(\d*)\]
        # 	_(\w*)$
        # ''',re.VERBOSE)
        nType = 0
        # print( 'filePath: ' + filePath + ' savePath: ' + savePath )
        data = xlrd.open_workbook(filePath)
        table = data.sheet_by_name(sheetname)
        commentInfo = []
        headerInfo = []
        xlsInfo = {}
        ids = []
        if x1 < 3 or x1 > table.nrows:
            x1 = 3
        if x2 < 2 or x2 > table.nrows:
            x2 = table.nrows

        for i in range(table.ncols):
            l = 2
            commentInfo.append(table.cell(0, i).value)
            header = table.cell(1, i).value
            headerInfo.append(header)
            match = regArr2.match(header)
            if match:
                nType = 3
            else:
                match = regArr1.match(header)
                if match:
                    nType = 2
                else:
                    match = reg.match(str(header))
                    # print(match.group(0))
                    nType = 1
            nId = 0
            for j in range(x1 - 1, x2):
                cellInfo = table.cell(j, i).value
                # 判断excel读取的元素的类型
                type = table.cell(j, i).ctype
                if str(cellInfo) == 'true':
                    cellInfo = True
                if str(cellInfo) == 'false':
                    cellInfo = False
                if type == 2 and cellInfo % 1 == 0:
                    cellInfo = int(cellInfo)
                elif type == 4:
                    cell = True if cell == 1 else False
                # end
                if i == 0:
                    # cellInfo = int(cellInfo)
                    ids.append(l - 1)
                    xlsInfo[l - 1] = {}
                    nType = 0
                nId = int(ids[l - 2])
                l += 1
                if nType == 0:
                    # if match.group(1) not in xlsInfo[nId] and reg.match(str(header))==False:
                    # 	xlsInfo[nId][match.group(1)] = {}
                    # 	xlsInfo[nId][match.group(1)][match.group(3)] = cellInfo
                    # if match.group(1) not in xlsInfo[nId] and reg.match(str(header))==False and regArr1.match(header):
                    if match.group(1) not in xlsInfo[nId] and regArr1.match(header):
                        xlsInfo[nId][match.group(1)] = {}
                        xlsInfo[nId][match.group(1)][match.group(3)] = cellInfo
                    else:
                        xlsInfo[nId][match.group(1)] = cellInfo
                if nType == 1:
                    xlsInfo[nId][match.group(1)] = cellInfo
                elif nType == 2:
                    if match.group(1) not in xlsInfo[nId]:
                        xlsInfo[nId][match.group(1)] = {}
                    xlsInfo[nId][match.group(1)][match.group(3)] = cellInfo
                elif nType == 3:
                    if match.group(1) not in xlsInfo[nId]:
                        xlsInfo[nId][match.group(1)] = []
                    if len(xlsInfo[nId][match.group(1)]) < int(match.group(2)):
                        xlsInfo[nId][match.group(1)].append({})
                    xlsInfo[nId][match.group(1)][int(match.group(2)) - 1][match.group(1)] = cellInfo

        # outFile = codecs.open( savePath, 'w', 'utf-8' )
        listinfo = []

        # for k in range(len(listinfo)):
        #	del listinfo[k]['ID']

        for key, value in list(xlsInfo.items()):
            listinfo.append(value)
        # print('序号:'+str(key)+'  json数据:'+str(value))

        return listinfo

    def split_data(self,value,fh=" "):
        '''
        切分数据,返回数组,例如:
        str=3.14.15
        |split data|str|
        return ['3','14','15']
        '''
        if not fh:
            fh=" ";
        return value.split(fh)

    def re_search(self,str,Ls,Rs):
        '''
        通过正则查询结果
        str 被切的数据
        Ls  左边界
        Rs  右边界
        如有多个只取第一个
        Examples:
        | re search | abcd | a | d | # 返回结果是bc
        '''
        m=re.search( Ls+'(.*?)'+Rs,str)
        if m is not None:
            return m.group(1)
            logger.debug('return'+m.group(1))
        else:
            logger.info(str)

    def re_search_all(self,str,Ls,Rs):
        '''
        通过正则查询结果
        str 被切的数据
        Ls  左边界
        Rs  右边界
        返回list
        Examples:
        | re search all | A111B  A222B | A | B | # 返回结果是['111','222']
        '''
        pat=re.compile(Ls+'(.*?)'+Rs)
        m=re.findall(pat,str)
        if m is not None:
            return m
        else:
            logger.info('re_search_all >> None')


    def Get_Time_Modified(self,addnumber='0'):
        '''
        获得当前日期. 可以通过参数加减日期
        :param addnumber: 加减天数, 默认是今天
        :return: str
        '''
        d1 = datetime.date.today()
        d2=d1+datetime.timedelta(int(addnumber))
        return d2

    def Get_Timestamp(self):
        '''
        获得时间戳
        :return: str , 保证数字唯一
        如: 1464921407
        '''
        res=time.time()
        return str(int(res))

    def Random_Num(self,start=1,stop=10000,times=1):
        '''
        随机产生一个随机数
        :param start 随机数最小值 默认是1
        :param stop  随机数最大值 默认是10000
        :param times 倍数,用于凑整随机, 默认是1
        :return: str
        如:
        Random Num | start=1 | stop=10 | times=100  返回 100 ~ 1000 的随机 返回结果为 100 或 200 等
        '''
        num=random.randint(int(start),int(stop))
        num=num*times
        logger.debug('生成随机数:'+str(num))
        return num

    def Random_Choice(self,sequence):
        '''
        随机选择有序类型(如数组)中的某一个值
        :param sequence 有序类型.
        :return 根据你传的参数决定类型
        如:
        Random Choice | ['a','b','c']  返回 a,b,c中的随机一个
        Random Choice | hello    返回h,e,l,l,o 中的随机一个
        '''
        res=random.choice(sequence)
        return res

    def json_Dumps(self,obj):
        '''
        :param obj: 字典或者str类型dumps后会变成json格式. 注意其他类型的会报错
        :return: json
        '''
        if type(obj) is str:
            obj=obj.encode('utf-8')
        logger.debug(type(obj))
        logger.debug(obj)
        if isinstance(obj,str):
            d=json.JSONDecoder().decode(obj)
            data=json.dumps(d)
        elif isinstance(obj,dict) or isinstance(obj,list):
            data=json.dumps(obj)
        else:
            logger.error("typeError: can't dumps "+str(type(obj)) +" . must <str> or <dict> ")
        return data

    def FormData_to_Dict(self,text):
        '''
        text格式参考 casenumber=&searoute=null&isExsitAdjunct=&currentDate=2016-02-05
        :param text: str
        :return:dict
        '''
        adict={}
        for a in text.split('&'):
            (key,value)= a.split('=')
            adict[key]=value
        return adict

    def Jsonstr_to_Dict(self,jsonStr):
        '''
        text格式参考json 如 {"a":1,"b":2,"3":"c","4":["k","k1"]}
        '''
        d=json.JSONDecoder().decode(jsonStr)
        return d
        
    def code_str(self,s,y):
        '''
        将enicode去掉U
        用逗号分割
        '''
        data=y.join(s)
        return data
        

    def dict_values(self,s):
        '''
        获取dictionary中的values值
        '''
        data=list(s.values())
        return data
        

    def steplog(self,msg):
        '''
        写入格式如:
        2015-12-14   XXXXX
        '''
        #print type(msg)
        #print msg
        #RF传入的是UnicodeType,先转成str
        if type(msg) is str:
            msg=msg.encode('utf-8')
        path=os.getcwd()
        projectpath=os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
        logpath=projectpath+os.sep+"steplog"
        if not os.path.exists(logpath):
            logpath=os.mkdir(projectpath+os.sep+"steplog")
        print(logpath)
        try:
            with open(logpath+os.sep+time.strftime("%Y-%m-%d")+'log.txt','a') as logs:
                logs.write(time.strftime("%H:%M:%S") + "    "+msg+"\n")
        except Exception as e:
            raise e


    def Get_advancedConditionsString(self,str):
        '''
        str demo : 起运港=NINGBO,目的港=DUBAI
        目前支持: 起运港 目的港
        '''

    def clear_host(self):
        '''
        清空文本中的内容
        说明：
        第一个参数是文件名称，包括路径如：C:\Windows\System32\drivers\etc\HOSTS；第二个参数是打开的模式mode
        'r'：只读（缺省。如果文件不存在，则抛出错误）
        'w'：只写（如果文件不存在，则自动创建文件）
        'a'：附加到文件末尾
        'r+'：读写
        '''
        print(os.path.isfile("C:\Windows\System32\drivers\etc\HOSTS"))


        file1 = open("C:\Windows\System32\drivers\etc\HOSTS","r+")

        file1.truncate()

        file1.close()

    def clear_document(self,str):
        '''
        清空文本中的内容
        str:文件所在目录
        '''
        file1 = open (str,"r+")
        file1.truncate()
        file1.close()

    def write_host(self,str2):
        '''
        内容写入文本
        参数说明：
        str:文本所在路径
        str1:打开的模式mode,'r'：只读（缺省。如果文件不存在，则抛出错误）
        'w'：只写（如果文件不存在，则自动创建文件）
        'a'：附加到文件末尾
        'r+'：读写
        str2:写入的内容
        '''
        file ("C:\Windows\System32\drivers\etc\HOSTS","r+").writelines(str2)
        file1 = open ("C:\Windows\System32\drivers\etc\HOSTS","r+")
        file2 = file1.readlines()
        print(file2)
        file1.close()

    def write_document(self,str,str1):
        '''
        写入文件内容
        str:文件所在路径
        str1:写入的内容
        '''

        print(os.path.isfile(str))
        file (str,"r+").writelines(str1)
        file1 = open(str,"r+")
        print(file1.readlines())
        file1.close()

    def read_document(self,str):
        '''
        读取文本中的多行
        str:文件所在路径
        '''
        file1 = open (str,"r+")
        file2 = file1.readlines()
        return file2
        file1.close()

    def set_download_dir(self,dir1,dir2):
    	'''
    	指定谷歌浏览器下载路径，但是下载动作需要自己添加click
    	dir1:下载的路径
    	dir2:浏览器驱动所在的路径
    	url:下载的链接
    	location:定位附件位置
    	eg: 在某个链接下载zip文件
    	file_download   dir1 = "d:\\"   dir2 ='D:\\python\\chromedriver.exe'  url = 'http://sahitest.com/demo/saveAs.htm'  location = '//a[text()="testsaveas.zip"]'
    	'''
    	options = webdriver.ChromeOptions()
    	prefs = {'profile.default_content_settings.popups': 0,'download.default_directory': dir1}
    	options.add_experimental_option('prefs',prefs)
    	options.add_argument('--user-agent=iphone')
    	driver = webdriver.Chrome(executable_path=dir2, chrome_options=options)
    	sleep (10)

    def clear_space_all(self,str):
    	'''
    	清空str的前后空格
    	'''
    	return str.strip()

    # def create_excel(self, filename, sheetname):
    #     '''
    #         在指定位置创建空白的excel文件
    #         filename:生成的新文件位置
    #         sheetname:生成的sheet名
    #     '''
    #     filename = filename.decode("utf-8")  # 将字符串文件编码改为unicode
    #
    #     workbook = xlwt.Workbook(encoding='utf-8')  # 新建工作簿
    #     sheet1 = workbook.add_sheet(sheetname)  # 新建sheet
    #     workbook.save(filename)  # 保存
    #
    # def add_sheet(self, filename, sheetname):
    #     '''
    #     在存在的Excel中新增sheet
    #     :param filename: 已存在的Excel名
    #     :param sheetname: 新增的sheet表名
    #     :return:
    #     '''
    #     filename = filename.decode("utf-8")  # 将字符串文件编码改为unicode
    #
    #     wb = open_workbook(filename, formatting_info=True)
    #     # 复制原有表
    #     newb = copy(wb)
    #     # 新增sheet,参数是该sheet的名字，可自定义
    #     wbsheet = newb.add_sheet(sheetname)
    #     newb.save(filename)
    #
    # def create_excel_if_noexist(self, filename, sheetname):
    #     '''
    #     当Excel文件或Sheet表不存在时创建文件
    #     :param filename: 文件路径
    #     :param sheetname: Sheet表名
    #     :return:
    #     '''
    #     filename = filename.decode("utf-8")  # 将字符串文件编码改为unicode
    #
    #     if os.path.exists(filename):
    #         excel = open_workbook(filename, formatting_info=True)
    #         sheetList = excel.sheet_names()  # 获取Excel的所有Sheet名
    #         if sheetname not in sheetList:
    #             self.add_sheet(filename, sheetname)
    #     else:
    #         self.create_excel(filename, sheetname)
    #
    # def set_font_style(self, color = 0, size = 200):
    #     '''
    #     设置单元格字体格式
    #     :param color: 设置Excel的数据字体颜色索引，默认为0；0 = 黑； 1 = 白； 2 = 红； 3 = 绿, 4 = 蓝, 5 = 黄
    #     :param size: 写入的Excel的数据字体大小，默认为200(10*20)
    #     :return:
    #     '''
    #     font = xlwt.Font()
    #     font.colour_index = color
    #     font.height = size
    #
    #     style = xlwt.XFStyle()
    #     style.font = font
    #
    #     return style
    #
    # def write_all_excel(self, filename, sheetname, data, color = 0, size = 200):
    #     '''
    #     将数据从头写入已存在的Excel文件
    #     :param filename: 写入的Excel路径
    #     :param sheetname: 写入的Excel的表名
    #     :param data: 写入的Excel数据
    #     :param color: 写入的Excel的数据字体颜色索引，默认为0；0 = 黑； 1 = 白； 2 = 红； 3 = 绿, 4 = 蓝, 5 = 黄
    #     :param size: 写入的Excel的数据字体大小，默认为200(10*20)
    #     :return:
    #     '''
    #     filename = filename.decode("utf-8")  # 将字符串文件编码改为unicode
    #
    #     color = color if isinstance(color, int) else int(color)
    #
    #     col_count = len(data) # 获取data这个list的元素个数
    #     style = self.set_font_style(color, size)  # 设置字体样式
    #     self.create_excel_if_noexist(filename, sheetname)  # 判断文件是否存在，否则就创建
    #
    #     rbook = open_workbook(filename, formatting_info=True)
    #     wb = copy(rbook)
    #     sheetIndex = rbook.sheet_names().index(sheetname)  # 获取sheet表所在的索引位置
    #     ws = wb.get_sheet(sheetIndex)
    #
    #     # 带格式循环写入Excel文件
    #     for row in range(col_count):
    #         for col in range(len(data[row])):
    #             # 如果是非int或float型的数据时转换编码格式
    #             cell_data = data[row][col] if (isinstance(data[row][col], int) or isinstance(data[row][col], float)) else data[row][col].decode("utf-8")
    #             ws.write(row, col, cell_data, style)
    #
    #     wb.save(filename)  # 保存Excel文件
    #
    #
    # def write_cell(self, filename, sheetname, row, col, data, color = 0, size = 200):
    #     '''
    #     带格式写入Excel某个单元格
    #     :param filename: Excel文件路径
    #     :param sheetname: Excel对应表的名称
    #     :param row: 写入的单元格行位置，第一行为0
    #     :param col: 写入的单元格列位置，第一列为0
    #     :param data: 写入的Excel数据
    #     :param color: 写入的Excel的数据字体颜色索引，默认为0；0 = 黑； 1 = 白； 2 = 红； 3 = 绿, 4 = 蓝, 5 = 黄
    #     :param size: 写入的Excel的数据字体大小，默认为200(10*20)
    #     :return:
    #     '''
    #     filename = filename.decode("utf-8")  # 将字符串文件编码改为unicode
    #
    #     color = color if isinstance(color, int) else int(color)
    #     row = row if isinstance(row, int) else int(row)  # 不是int型时取对应int值
    #     col = col if isinstance(col, int) else int(col)
    #
    #     style = self.set_font_style(color, size)  # 设置字体样式
    #     self.create_excel_if_noexist(filename, sheetname)  # 判断文件是否存在，否则就创建
    #
    #     # 带格式写入
    #     rbook = open_workbook(filename, formatting_info=True)
    #     wb = copy(rbook)
    #     sheetIndex = rbook.sheet_names().index(sheetname)
    #     ws = wb.get_sheet(sheetIndex)
    #     # 如果是非int或float型的数据时转换编码格式
    #     cell_data = data if (isinstance(data, int) or isinstance(data, float)) else data.decode("utf-8")
    #     ws.write(row, col, cell_data, style)
    #
    #     wb.save(filename)  # 保存Excel文件


    def buildSign(self, param, secret):
        keys = list(param.keys())  # 获取对应字典的key
        sort_keys = sorted(keys)  # 对key进行升序排序

        paramNameValue = ""
        for name in sort_keys:
            paramNameValue += name
            paramNameValue += param[name]

        source = secret + str(paramNameValue) + secret
        print(source)
        return self.md5(source)


    def md5(self, message):
        m = hashlib.md5()
        m.update(message.encode('utf-8'))
        str_md5 = m.digest()
        l = ['{:02X}'.format(i) for i in str_md5]      # ord():ord(单字符) 转换为ascii值
        s = ''.join(l)
        return s


    # def open_and_read_cell(self, filename, sheetname, cell_name):
    #     filename = filename.decode("utf-8")  # 将字符串文件编码改为unicode
    #     wb = xlrd.open_workbook(filename)
    #     sheetNames = wb.sheet_names()
    #     my_sheet_index = sheetNames.index(sheetname)
    #     sheet = wb.sheet_by_index(my_sheet_index)
    #     for row_index in range(sheet.nrows):
    #         for col_index in range(sheet.ncols):
    #             cell = cellname(row_index, col_index)
    #             if cell_name == cell:
    #                 cellValue = sheet.cell(row_index, col_index).value
    #     return cellValue


    def XlsToResultDic(self, filePath, sheetname='Sheet1', start_row=0, end_row=0, check_col = 0):
        start_row = int(start_row)
        end_row = int(end_row)
        check_col = int(check_col)
        reg = re.compile('^(\w*)$')
        regArr1 = re.compile(r'''
    		^(\w*)
    		\[(\d*)\]
    		_(\w*)$
    		''', re.VERBOSE)
        regArr2 = re.compile(r'''
    		^(\w*)
    		\[(\d*)\]
    		<(\w*)>$''', re.VERBOSE)
        nType = 0
        data = xlrd.open_workbook(filePath)
        table = data.sheet_by_name(sheetname)
        commentInfo = []
        headerInfo = []
        xlsInfo = {}
        ids = []

        # 循环读取数据
        for i in range(check_col):
            l = 2
            commentInfo.append(table.cell(0, i).value)
            header = table.cell(1, i).value
            headerInfo.append(header)
            match = regArr2.match(header)
            if match:
                nType = 3
            else:
                match = regArr1.match(header)
                if match:
                    nType = 2
                else:
                    match = reg.match(str(header))
                    nType = 1
            nId = 0
            for j in range(start_row - 1, end_row):
                cellInfo = table.cell(j, i).value
                # 判断excel读取的元素的类型
                type = table.cell(j, i).ctype
                if str(cellInfo) == 'true':
                    cellInfo = True
                if str(cellInfo) == 'false':
                    cellInfo = False
                if type == 2 and cellInfo % 1 == 0:
                    cellInfo = int(cellInfo)
                elif type == 4:
                    cellInfo = True if cellInfo == 1 else False
                # end
                if i == 0:
                    ids.append(l - 1)
                    xlsInfo[l - 1] = {}
                    nType = 0
                nId = int(ids[l - 2])
                l += 1
                if nType == 0:
                    if match.group(1) not in xlsInfo[nId] and regArr1.match(header):
                        xlsInfo[nId][match.group(1)] = {}
                        xlsInfo[nId][match.group(1)][match.group(3)] = cellInfo
                    else:
                        xlsInfo[nId][match.group(1)] = cellInfo
                if nType == 1:
                    xlsInfo[nId][match.group(1)] = cellInfo
                elif nType == 2:
                    if match.group(1) not in xlsInfo[nId]:
                        xlsInfo[nId][match.group(1)] = {}
                    xlsInfo[nId][match.group(1)][match.group(3)] = cellInfo
                elif nType == 3:
                    if match.group(1) not in xlsInfo[nId]:
                        xlsInfo[nId][match.group(1)] = []
                    if len(xlsInfo[nId][match.group(1)]) < int(match.group(2)):
                        xlsInfo[nId][match.group(1)].append({})
                    xlsInfo[nId][match.group(1)][int(match.group(2)) - 1][match.group(1)] = cellInfo

        listinfo = []
        for key, value in list(xlsInfo.items()):
            iteminfo = []
            iteminfo.append(value)
            listinfo.append(iteminfo)

        # 拼接验证数据
        checkHeaderInfo = []
        checkInfo = {}
        checkIds = []
        checkItem = {}
        for col in range(check_col, table.ncols):
            l = 2
            checkHeader = table.cell(1, col).value
            checkHeaderInfo.append(checkHeader)
            checkKey = checkHeader.decode("utf-8")

            for row in range(start_row - 1, end_row):
                if col == check_col:
                    checkIds.append(l - 1)
                    checkInfo[l - 1] = []
                checknId = int(checkIds[l - 2])
                print(checknId)
                l += 1
                cellInfo = table.cell(row, col).value
                type = table.cell(row, col).ctype
                if str(cellInfo) == 'true':
                    cellInfo = True
                if str(cellInfo) == 'false':
                    cellInfo = False
                if type == 2 and cellInfo % 1 == 0:
                    cellInfo = int(cellInfo)

                checkInfo[checknId].append(cellInfo)

        checkList = []
        for key, value in list(checkInfo.items()):
            listinfo[key - 1].append(value)
        return listinfo
        # 
        # resultDic = {}
        # resultDic["data"] = listinfo
        # resultDic["check"] = checkList

        # return resultDic

    # 以下excel方法基于python3编写
    def create_excel(self, filename, sheetname):
        '''
        创建excel
        :param filename: 创建excel的文件名
        :param sheetname: 需要的sheet名
        :return:
        '''
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(title=sheetname, index=0)
        wb.save(filename)

    def add_sheet(self, filename, sheetname):
        '''
        在存在的Excel中新增sheet
        :param filename: 已存在的Excel名
        :param sheetname: 新增的sheet表名
        :return:
        '''
        wb = openpyxl.load_workbook(filename)
        wb.create_sheet(title=sheetname)
        wb.save(filename)

    def create_excel_if_noexist(self, filename, sheetname):
        '''
        判断excel是否存在，如果不存在的话新增一个有对应sheet名的excel
        :param filename: 需要维护的excel名
        :param sheetname: 需要新增的sheet名
        :return:
        '''
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)  # 存在时打开excel
            ws = wb.sheetnames  # 获取excel中所有的sheet名
            if sheetname not in ws:  # 如果不存在需要的sheet名时新增一个sheet
                wb.create_sheet(title=sheetname)
                wb.save(filename)
        else:
            self.create_excel(filename, sheetname)  # 若不存在excl时则创建一个

    def read_excel_all_data(self, filename, sheetname):
        '''
        读取excel对应sheet中的所有内容
        :param filename: 需读取的excel名
        :param sheetname: 需读取的sheet名
        :return:
        '''
        wb = openpyxl.load_workbook(filename)  # 读取excel文件
        ws = wb[sheetname]  # 根据sheet名读取excel
        max_row = ws.max_row  # 获取sheet表格的最大行数
        max_col = ws.max_column  # 获取sheet表格的最大列数
        data = []
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                row_data.append(ws.cell(row, col).value)
            data.append(row_data)

        return data

    def read_cell_by_index(self, filename, sheetname, row, col):
        """
        根据行列号读取单元格内容(如读取B4，则row=4, column=2)
        :param filename: 需读取的excel名
        :param sheetname: 需读取的sheet名
        :param row: 需读取的行号
        :param col: 需读取的列名
        :return:
        """
        wb = openpyxl.load_workbook(filename)  # 读取excel文件
        ws = wb[sheetname]  # 根据sheet名读取excel
        cell_vaule = ws.cell(row, col).value
        return cell_vaule

    def read_cell_by_name(self, filename, sheetname, name):
        """
        根据行列号读取单元格内容
        :param filename: 需读取的excel名
        :param sheetname: 需读取的sheet名
        :param name: 需读取的单元格名
        :return:
        """
        wb = openpyxl.load_workbook(filename)  # 读取excel文件
        ws = wb[sheetname]  # 根据sheet名读取excel
        cell_vaule = ws[name].value
        return cell_vaule

    def set_font_style(self, color, size, bold=False):
        '''
        设置字体大小和颜色
        :param color: 6位色号，如000000
        :param size:  字的大小，数字
        :param bold:  是否加粗，默认为False
        :return:
        '''
        size = size if isinstance(size, int) else int(size)
        font = Font(size=size, bold=bold, color=color)
        return font

    def write_all_excel(self, filename, sheetname, data, color=u'000000', size=11, bold=False):
        '''
        写入excel
        :param filename:  需写入的excel文件名
        :param sheetname:  需写入的sheet名
        :param data:  写入的内容
        :param color:  6位色号，默认为000000
        :param size:  字的大小，数字，默认为11
        :param bold:  是否加粗，默认为False
        :return:
        '''
        self.create_excel_if_noexist(filename, sheetname)

        max_row = len(data)
        font = self.set_font_style(color, size, bold)

        wb = openpyxl.load_workbook(filename)
        ws = wb[sheetname]
        for row in range(max_row):
            for col in range(len(data[row])):
                ws.cell(row + 1, col + 1, data[row][col]).font = font

        wb.save(filename)

    def write_cell(self, filename, sheetname, row, col, data, color=u'000000', size=11, bold=False):
        '''
        写入excel
        :param filename:  需要写入的excel名
        :param sheetname:  需要写入的sheet名
        :param row:  需要写入的行
        :param col:  需要写入的列
        :param data:  需要写入的内容
        :param color:  6位色号，默认为000000
        :param size:  字的大小，数字，默认为11
        :param bold:  是否加粗，默认为False
        :return:
        '''
        row = row if isinstance(row, int) else int(row)
        col = col if isinstance(col, int) else int(col)

        font = self.set_font_style(color, size, bold)

        wb = openpyxl.load_workbook(filename)
        ws = wb[sheetname]
        ws.cell(row, col, data).font = font
        wb.save(filename)


if __name__ == '__main__':
    filename = "C:\\Users\\yusen\\Desktop\\result_inter_test.xlsx"
    sheetname = "运价--海运出口整箱"
    data = [\
        ["NO.",	"Api Purpose",	"Protocol",	"Api Host",	"Request Url",	"Request Method",	"Request Data Type", "Request Data",	"Check Point"],\
        [1, "查询获取状态", "Http", "gm1-scysa.100jit.com", "/fms-rest/rest/baseGoodsStatus/select", "GET", "Form", "searchValue=", "\"resultCode\":100"],\
            ]
    olym = OlymKeywords()
    olym.create_excel(filename, sheetname)
    olym.write_all_excel(filename, sheetname, data)
    olym.write_cell(filename, sheetname, 1, 10, "新写入的字段", color = 'ff0000')
    a = olym.read_cell_by_name(filename, sheetname, u'J1')
    print(a)
    olym.write_cell(filename, sheetname, 2, 1, 11.331, size=16)
    b = olym.read_cell_by_name(filename, sheetname, u'A2')
    print(b)

    # {'name': 'booking.mains.callback', 'version': '1.0', 'app_key': 'yjt_test',
    #  'data': "{'casenumber':'MTHREE2714962','status':-10,'rejuectReason':'拒绝你还需要理由？'}",
    #  'timestamp': '2019-08-23 11:16:09', 'format': 'json'}

    # appKey = "yjt_test"
    # secret = "123456"
    # param = {}
    # param['name'] = "booking.mains.callback"
    # param['version'] = "1.0"
    # param['app_key'] = appKey
    # param['timestamp'] = "2019-08-23 11:16:09"  # timestamp=2019-03-06 17:25:42
    # param['format'] = "json"
    # param['data'] = "%7B%27casenumber%27%3A%27MTHREE2714962%27%2C%27status%27%3A-10%2C%27rejuectReason%27%3A%27%E6%8B%92%E7%BB%9D%E4%BD%A0%E8%BF%98%E9%9C%80%E8%A6%81%E7%90%86%E7%94%B1%EF%BC%9F%27%7D"
    # #
    # # param =  {'name': 'booking.mains.callback', 'version': '1.0', 'app_key': 'yjt_test', 'data': '%7B%27casenumber%27%3A%27MTHREE2714962%27%2C%27status%27%3A-10%2C%27rejuectReason%27%3A%27%E6%8B%92%E7%BB%9D%E4%BD%A0%E8%BF%98%E9%9C%80%E8%A6%81%E7%90%86%E7%94%B1%EF%BC%9F%27%7D', 'timestamp': '2019-08-23 15:11:24', 'format': 'json'}
    # olym = OlymKeywords()
    # sign = olym.buildSign(param, secret)
    # print(sign)
    # filename = "C:\\Users\\yusen\\Desktop\\test测试.xls"
    # olym = OlymKeywords()
    # a = olym.XlsToResultDic(filename, "haha", 3, 4, 16)
    # print(a)

    
    
