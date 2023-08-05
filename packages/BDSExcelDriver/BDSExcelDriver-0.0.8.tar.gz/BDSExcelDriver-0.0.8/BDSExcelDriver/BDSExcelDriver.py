from openpyxl import workbook
from openpyxl import load_workbook
from robot.api import logger
from SeleniumLibrary.base import keyword, LibraryComponent
from robot.libraries.BuiltIn import BuiltIn
import json


class BDSExcelDriver:
    def __init__(self, file=None):
        self.file = file

    @keyword
    def _write_to_existing_excel(self, value, column='A', row='1', sheet_name='Sheet1'):
        wb = load_workbook(self.file)
        if(wb.sheetnames.count(sheet_name) == 0):
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
        sheet[column + row] = value
        wb.save(self.file)

    @keyword
    def _read_from_existing_excel(self, column='A', row='1', sheet_name='Sheet1'):
        wb = load_workbook(self.file)
        if(wb.sheetnames.count(sheet_name) == 0):
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
        result = sheet[column + row].value
        if(result == '${EMPTY}'):
            result = ''
        return result
